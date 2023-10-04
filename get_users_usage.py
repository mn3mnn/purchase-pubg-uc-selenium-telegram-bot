import sqlite3
import openpyxl


def get_data_from_db():
    conn = sqlite3.connect('pubg_uc.db')
    c = conn.cursor()
    c.execute('SELECT user, name, mobile, count(*) FROM codes_users LEFT JOIN users ON users.username = codes_users.user GROUP BY user')
    data = c.fetchall()
    conn.close()
    return data


def write_data_to_excel():
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        ws['A1'] = 'Username'
        ws['B1'] = 'Name'
        ws['C1'] = 'Mobile'
        ws['D1'] = 'Usage'
        data = get_data_from_db()
        for row in data:
            ws.append(row)

        wb.save('usage.xlsx')
        print('usage.xlsx file created successfully.')

    except Exception as e:
        print(e)


if __name__ == '__main__':
    write_data_to_excel()

